from copy import copy
from openpyxl import load_workbook
from .exstyles import (
    ummanstyle,
    BoldFont,
    sdwstyle,
    tscetstyle,
    aligncenter,
    verticalcenter,
    text_red,
    standart_font,
    unvisible,
    small_font,
    balancestyle,
)
from openpyxl.styles import Font, Color, Alignment, Border, Side, NamedStyle
from openpyxl.formula.translate import Translator
from openpyxl.styles import Border, Side
import os
from .settings import MEDIA_ROOT


class Exlsx:
    def __init__(self, operations, start_data, template_file):
        self.path = "Balance.xlsx"
        self.template_path = os.path.join(MEDIA_ROOT, "Template.xlsx")
        self.wb = load_workbook(self.template_path)
        self.operations = operations  # [{}]
        self.start_data = start_data  # [{}]
        self.spisok_file = template_file  # spisok.xlsx
        # self.spisok_file = os.path.join(MEDIA_ROOT, "spisok.xlsx")  # spisok.xlsx

    def main(self):
        spisokrange = self.copyspisok()
        self.handleoperations()
        unique_list, sdwrow, sdw = self.handlesdw(spisokrange)
        self.handletscet(unique_list, sdw)
        countrow = self.handlenulevoi(sdwrow, sdw)
        self.handlebalans(sdwrow, countrow)
        return self.wb

    # SPISOK
    def copyspisok(self):
        # SPISOK
        wb_source = load_workbook(self.spisok_file)
        ws_source = wb_source["Список счетов"]
        ws_target = self.wb["Список счетов"]
        mr = self.copy_shet(ws_source, ws_target)
        spisokrange = mr
        return spisokrange

    def copy_shet(self, ws_source, ws_target):

        mr = ws_source.max_row
        mc = ws_source.max_column

        # Unmerge all cells
        for merge in list(ws_source.merged_cells):
            ws_source.unmerge_cells(range_string=str(merge))

        for merge in list(ws_target.merged_cells):
            ws_target.unmerge_cells(range_string=str(merge))

        # # copying the cell values from source
        # excel file to destination excel file
        for i in range(1, mr + 1):
            for j in range(1, mc + 1):
                # reading cell value from source excel file
                c = ws_source.cell(row=i, column=j)

                # writing the read value to destination excel file
                ws_target.cell(row=i, column=j).value = c.value

        # Merge only D AND E columns
        for row in range(1, mr):
            dcell = ws_target.cell(row, column=4)
            if dcell.value == None:
                # print("NO DCELL VALUE", dcell.value)
                ecell = ws_target.cell(row, column=5)
                dcell.value = ecell.value
            ws_target.merge_cells(f"D{row}:E{row}")

        return mr

    # OPERATIONS
    def handleoperations(self):
        ws = self.wb["Операциялар"]
        self.operationsFn(ws, self.operations)
        print("PASSED OPERATIONS")

    def operationsFn(self, ws, operations):
        # Операциялар
        count = 1
        for op in operations:
            # Applying values
            row = 3 + count
            # print(row)
            ws.cell(row, column=2).value = count
            ws.cell(row, column=3).value = op["name"]
            ws.cell(row, column=4).value = op["Do"]
            ws.cell(row, column=5).value = op["Ko"]
            ws.cell(row, column=6).value = op["sum"]
            ws.cell(row, column=7).value = op["time"]
            # Applying styles
            ws.row_dimensions[row].height = 25
            row = ws[row]
            Acolumn = ws["A"]
            DEColumns = ws["D:E"]
            CColumn = ws["C"]
            for cell in row:
                if cell in Acolumn:
                    pass
                    # print("In A column", cell)
                else:
                    cell.style = ummanstyle
                    # print(cell)
                    if cell in CColumn:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    if cell in DEColumns[0] or cell in DEColumns[1]:
                        # print("In Decolumn", cell)
                        cell.font = BoldFont
            count += 1

    # SDW
    def handlesdw(self, spisokrange):
        # SDW
        sdw = self.wb["SDW"]
        unique_list, sdwrow = self.sdwFn(
            sdw=sdw,
            operations=self.operations,
            startData=self.start_data,
            spisokrange=spisokrange,
        )
        print("PASSED SDW")
        return unique_list, sdwrow, sdw

    def duplicateRow(self, sdw, oldRow, newRow=3):
        for cell in oldRow:
            new_cell = sdw.cell(row=newRow, column=cell.col_idx, value=cell.value)
            # print("\nFormula", cell.value)
            # print("Coordinate", cell.coordinate)
            val = str(cell.value)
            if val and val.startswith("="):
                new_cell.value = Translator(
                    cell.value, cell.coordinate
                ).translate_formula(new_cell.coordinate)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    def sdwFn(self, sdw, operations, startData, spisokrange):
        count = 1
        unique_list = []
        unique_scets = []

        #  function to get unique scets
        def unique(list1):
            # traverse for all elements
            for x in list1:
                # check if exists in unique_list or not
                if "cs" in x:
                    if x["cs"] not in unique_scets:
                        unique_list.append({"cs": x["cs"], "cn": x["sum"]})
                        unique_scets.append(x["cs"])
                        # print("APPENDED: ", x["cs"])
                if "Do" in x:
                    if x["Do"] not in unique_scets:
                        unique_list.append({"cs": x["Do"], "cn": 0})
                        unique_scets.append(x["Do"])
                        # print("APPENDED: ", x["Do"])
                if "Ko" in x:
                    if x["Ko"] not in unique_scets:
                        unique_list.append({"cs": x["Ko"], "cn": 0})
                        unique_scets.append(x["Ko"])
                        # print("APPENDED: ", x["Ko"])

        unique(startData)
        unique(operations)

        # TOP Add SCETS
        for cset in unique_list:
            row = 4 + count
            # first time 23
            # second time row-1
            if count == 1:
                self.duplicateRow(sdw, oldRow=sdw[23], newRow=row)
                sdw.delete_rows(23)
            else:
                self.duplicateRow(sdw, oldRow=sdw[row - 1], newRow=row)
            if count == 1:
                sdw.cell(row, column=2).value = count
            else:
                sdw.cell(row, column=2).value = f"=B{row-1}+1"
            sdw.cell(row, column=3).value = cset["cs"]
            sdw.cell(
                row, column=4
            ).value = (
                f"=INDEX('Список счетов'!$B$1:$B${spisokrange},MATCH(C{row},список,0))"
            )
            sdw.cell(
                row, column=5
            ).value = (
                f"=INDEX('Список счетов'!$A$1:$A${spisokrange},MATCH(C{row},список,0))"
            )
            sdw.cell(
                row, column=6
            ).value = (
                f"=INDEX('Список счетов'!$D$1:$D${spisokrange},MATCH(C{row},список,0))"
            )
            # print("FORMULAS IN D column SDW", sdw.cell(row, column=4).value)
            sdw.cell(row, column=7).value = f'=IF(D{row}="А",{cset["cn"]},0)'
            sdw.cell(row, column=8).value = f'=IF(D{row}="П",{cset["cn"]},0)'

            # Apply styles TOP
            sdw.row_dimensions[row].height = 30
            count += 1

        # BOTTOM result
        row += 1
        sdw.merge_cells(start_row=row, end_row=row, start_column=3, end_column=6)
        sdw.cell(row, column=3, value="Б А Л А Н С")
        sdw.cell(row, column=7).value = f"=SUM($G$5:G{row-1})"
        sdw.cell(row, column=8).value = f"=SUM($H$5:H{row-1})"
        sdw.cell(row, column=9).value = f"=SUM($I$5:I{row-1})"
        sdw.cell(row, column=10).value = f"=SUM($J$5:J{row-1})"
        sdw.cell(row, column=11).value = f"=SUM($K$5:K{row-1})"
        sdw.cell(row, column=12).value = f"=SUM($L$5:L{row-1})"
        # BOTTOM Apply styles
        for cell in sdw[row]:
            data = [sdw.cell(row, column=i) for i in range(1, 13)]
            if cell in data:
                cell.style = sdwstyle
            if cell in sdw["C"]:
                cell.font = Font(name="Times New Roman", size=16, bold=True)
        sdw.row_dimensions[row].height = 30
        # Remove A column borders
        for cell in sdw["A"]:
            cell.border = Border(
                bottom=Side(border_style=None),
                top=Side(border_style=None),
                left=Side(border_style=None),
            )

        return unique_list, row

    # TSCET
    def handletscet(self, unique_list, sdw):
        # Т счет
        tscet = self.wb["Т-счёт"]
        self.tscetFn(tscet, unique_list, sdw)
        print("PASSED T-SCET")

    def change_lines_t(self, opSum, new_ws, row, xcount):
        # A9 - Date Do
        cell = new_ws.cell(row, column=1)
        formula = f'=IFERROR(INDEX(Операциялар!$B$4:$G${opSum},SMALL(IF($C$4=Операциялар!$D$4:$D${opSum},ROW(Операциялар!$G$4:$G${opSum})-3),ROW(D{xcount})),6),"")'
        cell.value = formula
        new_ws.formula_attributes[cell.coordinate] = {
            "t": "array",
            "ref": f"{cell.coordinate}:{cell.coordinate}",
        }

        # B9 - Kor scet Do
        cell = new_ws.cell(row, column=2)
        formula = f'=IFERROR(INDEX(Операциялар!$B$4:$F${opSum},SMALL(IF($C$4=Операциялар!$D$4:$D${opSum},ROW(Операциялар!$E$4:$E${opSum})-3),ROW(B{xcount})),4),"")'
        cell.value = formula
        new_ws.formula_attributes[cell.coordinate] = {
            "t": "array",
            "ref": f"{cell.coordinate}:{cell.coordinate}",
        }
        # C9 - Count Do
        cell = new_ws.cell(row, column=3)
        formula = f'=IFERROR(INDEX(Операциялар!$B$4:$F${opSum},SMALL(IF($C$4=Операциялар!$D$4:$D${opSum},ROW(Операциялар!$B$4:$B${opSum})-3),ROW(A{xcount})),1),"")'
        cell.value = formula
        new_ws.formula_attributes[cell.coordinate] = {
            "t": "array",
            "ref": f"{cell.coordinate}:{cell.coordinate}",
        }
        # D9 - Sum Do
        cell = new_ws.cell(row, column=4)
        formula = f'=IFERROR(INDEX(Операциялар!$D$4:$F${opSum},SMALL(IF($C$4=Операциялар!$D$4:$D${opSum},ROW(Операциялар!$F$4:$F${opSum})-3),ROW(A{xcount})),3),"")'
        cell.value = formula
        new_ws.formula_attributes[cell.coordinate] = {
            "t": "array",
            "ref": f"{cell.coordinate}:{cell.coordinate}",
        }
        # E9 - Sum Ko
        cell = new_ws.cell(row, column=5)
        formula = f'=IFERROR(INDEX(Операциялар!$D$4:$F${opSum},SMALL(IF($C$4=Операциялар!$E$4:$E${opSum},ROW(Операциялар!$F$4:$F${opSum})-3),ROW(B{xcount})),3),"")'
        cell.value = formula
        new_ws.formula_attributes[cell.coordinate] = {
            "t": "array",
            "ref": f"{cell.coordinate}:{cell.coordinate}",
        }
        # F9 - Count Ko
        cell = new_ws.cell(row, column=6)
        formula = f'=IFERROR(INDEX(Операциялар!$B$4:$F${opSum},SMALL(IF($C$4=Операциялар!$E$4:$E${opSum},ROW(Операциялар!$B$4:$B${opSum})-3),ROW(D{xcount})),1),"")'
        cell.value = formula
        new_ws.formula_attributes[cell.coordinate] = {
            "t": "array",
            "ref": f"{cell.coordinate}:{cell.coordinate}",
        }
        # G9 - Kor Ko
        cell = new_ws.cell(row, column=7)
        formula = f'=IFERROR(INDEX(Операциялар!$B$4:$F${opSum},SMALL(IF($C$4=Операциялар!$E$4:$E${opSum},ROW(Операциялар!$D$4:$D${opSum})-3),ROW(D{xcount})),3),"")'
        cell.value = formula
        new_ws.formula_attributes[cell.coordinate] = {
            "t": "array",
            "ref": f"{cell.coordinate}:{cell.coordinate}",
        }
        # H9 - Date Ko
        cell = new_ws.cell(row, column=8)
        formula = f'=IFERROR(INDEX(Операциялар!$B$4:$G${opSum},SMALL(IF($C$4=Операциялар!$E$4:$E${opSum},ROW(Операциялар!$G$4:$G${opSum})-3),ROW(G{xcount})),6),"")'
        cell.value = formula
        new_ws.formula_attributes[cell.coordinate] = {
            "t": "array",
            "ref": f"{cell.coordinate}:{cell.coordinate}",
        }

    def tscetFn(self, tscet, unique_list, sdw):
        count = 1
        opSum = len(self.operations) + 4
        for scet in unique_list:
            new_ws = self.wb.copy_worksheet(tscet)
            title = f"{count}-{scet['cs']}"
            new_ws.title = title

            # COPY TEMPLATE > CREATE NEW PAGE > PASTE FORMULAS

            # TOP
            new_ws["A3"] = count
            # SUM of rows in sdw list
            sum_row = len(unique_list) + 5

            new_ws["C3"] = f"=VLOOKUP(A3,SDW!$B$5:F{sum_row},5,0)"
            new_ws["C4"] = f"=VLOOKUP(A3,SDW!$B$5:F{sum_row},2,0)"
            new_ws["C5"] = f"=VLOOKUP(A3,SDW!$B$5:F{sum_row},3,0)"
            new_ws["C6"] = f"=VLOOKUP(A3,SDW!$B$5:F{sum_row},4,0)"
            new_ws["D8"] = f'=IF(C5="А",VLOOKUP(A3,SDW!$B$5:$H${sum_row},6,0),"")'
            new_ws["F8"] = f'=IF(C5="П",VLOOKUP(A3,SDW!$B$5:$H${sum_row},7,0),"")'

            ko = 0
            do = 0
            scet_count = 0
            # MIDDLE
            # Get all operations count for this scet to paste rows
            for ss in self.operations:
                if ss["Do"] == scet["cs"]:
                    scet_count += 1
                    do += 1
                elif ss["Ko"] == scet["cs"]:
                    scet_count += 1
                    ko += 1

            tscetrowcount = 9
            # If no operation for this scet do nothing
            if scet_count == 0:
                pass
            # If exist duplicate row upper from template
            else:
                if do >= ko:
                    for x in range(1, do):
                        tscetrowcount = 9 + x
                        self.duplicateRow(
                            new_ws, oldRow=new_ws[9], newRow=tscetrowcount
                        )
                else:
                    for x in range(1, ko):
                        tscetrowcount = 9 + x
                        self.duplicateRow(
                            new_ws, oldRow=new_ws[9], newRow=tscetrowcount
                        )
            xcount = 1
            # Empty template row if no operation
            if tscetrowcount - 9 == 0:
                self.change_lines_t(opSum, new_ws, 9, xcount)
                # OBOROTS IN SDW
                sdw.cell(4 + count, column=9).value = f"='{title}'!D10"
                sdw.cell(4 + count, column=10).value = f"='{title}'!F10"
                # KONECNYE IN SDW
                sdw.cell(
                    4 + count, column=11
                ).value = f"=IF(ISNUMBER('{title}'!$D11),'{title}'!$D11,0)"
                sdw.cell(
                    4 + count, column=12
                ).value = f"=IF(ISNUMBER('{title}'!$F11),'{title}'!$F11,0)"
            # Else add formulas for each line
            else:
                for row in range(9, tscetrowcount + 1):
                    self.change_lines_t(opSum, new_ws, row, xcount)
                    xcount += 1
                row += 1
                # Add Bottom
                new_ws.cell(row, column=3).value = "До="
                new_ws.cell(row, column=5).value = "Ко="
                new_ws.cell(row, column=4).value = f"=SUM(D9:D{row-1})"
                new_ws.cell(row, column=6).value = f"=SUM(E9:E{row-1})"
                # OBOROTS IN SDW
                sdw.cell(4 + count, column=9).value = f"='{title}'!D{row}"
                sdw.cell(4 + count, column=10).value = f"='{title}'!F{row}"

                row += 1
                new_ws.cell(row, column=3).value = f'=IF(C5="А","Ск=","")'
                new_ws.cell(row, column=5).value = f'=IF(C5="П","Ск=","")'
                new_ws.cell(
                    row, column=4
                ).value = f'=IF(C5="А",D8+D{row-1}-F{row-1},"")'
                new_ws.cell(
                    row, column=6
                ).value = f'=IF(C5="П",F8+F{row-1}-D{row-1},"")'
                # KONECNYE IN SDW
                sdw.cell(
                    4 + count, column=11
                ).value = f"=IF(ISNUMBER('{title}'!D{row}),'{title}'!D{row},0)"
                sdw.cell(
                    4 + count, column=12
                ).value = f"=IF(ISNUMBER('{title}'!F{row}),'{title}'!F{row},0)"
                # sdw.cell(4 + count, column=11).value = f"='{title}'!D{row}"
                # sdw.cell(4 + count, column=12).value = f"='{title}'!F{row}"

                # APPLY STYLES FOR BOTTOM
                row -= 1
                for x in range(3, 7):
                    new_ws.cell(row, column=x).style = tscetstyle
                new_ws.cell(row, column=4).border = Border(
                    bottom=Side(border_style="thin"),
                    top=Side(border_style="thin"),
                    right=Side(border_style="thin"),
                )
                new_ws.cell(row, column=4).alignment = aligncenter
                new_ws.cell(row, column=6).alignment = aligncenter
                new_ws.cell(row, column=3).alignment = verticalcenter
                new_ws.cell(row, column=5).alignment = verticalcenter
                new_ws.row_dimensions[row].height = 30
                row += 1
                new_ws.row_dimensions[row].height = 30
                for x in range(3, 7):
                    new_ws.cell(row, column=x).font = standart_font
                new_ws.cell(row, column=3).alignment = verticalcenter
                new_ws.cell(row, column=5).alignment = verticalcenter
                new_ws.cell(row, column=4).alignment = aligncenter
                new_ws.cell(row, column=4).font = text_red
                new_ws.cell(row, column=6).alignment = aligncenter
                new_ws.cell(row, column=6).font = text_red

            count += 1
        self.wb.remove(tscet)

    # NULEVOI
    def handlenulevoi(self, sdwrow, sdw):
        sdwrow += 20
        sdw_sum_row = sdwrow - 21
        countrow = sdwrow
        for x in range(sdw_sum_row):
            # UNIQUE balance types
            cell = sdw.cell(row=countrow, column=5)
            formula = f'=IFERROR(INDEX($E$5:$E${sdw_sum_row}, MATCH(0, INDEX(COUNTIF(E${sdwrow-2}:E{countrow-1}, $E$5:$E${sdw_sum_row}), 0, 0), 0)), "")'
            cell.value = formula
            cell.font = unvisible

            # CH - Do
            cell = sdw.cell(row=countrow, column=6)
            formula = f"=SUMIF($E$5:$E${sdw_sum_row},E{countrow},$G$5:$G${sdw_sum_row})"
            cell.value = formula
            cell.font = unvisible

            # CH - Ko
            cell = sdw.cell(row=countrow, column=7)
            formula = f"=SUMIF($E$5:$E${sdw_sum_row},E{countrow},$H$5:$H${sdw_sum_row})"
            cell.value = formula
            cell.font = unvisible

            # CH - Sum
            cell = sdw.cell(row=countrow, column=8)
            formula = f"=SUM(F{countrow}+G{countrow})"
            cell.value = formula
            cell.font = unvisible

            # CK - Do
            cell = sdw.cell(row=countrow, column=9)
            formula = f"=SUMIF($E$5:$E${sdw_sum_row},E{countrow},$K$5:$K${sdw_sum_row})"
            cell.value = formula
            cell.font = unvisible

            # CK - Ko
            cell = sdw.cell(row=countrow, column=10)
            formula = f"=SUMIF($E$5:$E${sdw_sum_row},E{countrow},$L$5:$L${sdw_sum_row})"
            cell.value = formula
            cell.font = unvisible

            # CK - Sum
            cell = sdw.cell(row=countrow, column=11)
            formula = f"=SUM(I{countrow}+J{countrow})"
            cell.value = formula
            cell.font = unvisible

            countrow += 1
        print("PASSED NULEVOI", sdwrow)
        return countrow

    # BALANS
    def handlebalans(self, sdwrow, countrow):
        # BALANS
        balans_ws = self.wb["Баланс"]

        exceptlist = [
            8,
            9,
            18,
            19,
            27,
            28,
            29,
            30,
            33,
            34,
            41,
            42,
            43,
            44,
            45,
            46,
            58,
            59,
            67,
            68,
            75,
            76,
            77,
        ]
        for x in range(8, 78):
            # APPLY BORDERS
            for column in range(4, 7):
                balans_ws.cell(row=x, column=column).style = balancestyle
            if x in exceptlist:
                for column in range(4, 7):
                    balans_ws.cell(row=x, column=column).font = small_font
            else:
                # CK
                cell = balans_ws.cell(row=x, column=5)
                cell.value = f"=L{x}"
                # cell.font = unvisible

                # CH
                cell = balans_ws.cell(row=x, column=6)
                cell.value = f"=P{x}"
                # cell.font = unvisible

                # 10 - CK +
                cell = balans_ws.cell(row=x, column=10)
                cell.value = (
                    f"=IFERROR(VLOOKUP(C{x},SDW!E${sdwrow}:$K${countrow},7,0),0)"
                )
                cell.font = unvisible

                # 11 - CK -
                cell = balans_ws.cell(row=x, column=11)
                cell.value = (
                    f"=IFERROR(VLOOKUP(H{x},SDW!E${sdwrow}:$K${countrow},7,0),0)"
                )
                cell.font = unvisible

                # 12 - CK =
                cell = balans_ws.cell(row=x, column=12)
                cell.value = f'=IFERROR(J{x}-K{x},"")'
                cell.font = unvisible

                # 14 - CH +
                cell = balans_ws.cell(row=x, column=14)
                cell.value = (
                    f"=IFERROR(VLOOKUP(C{x},SDW!E${sdwrow}:$K${countrow},4,0),0)"
                )
                cell.font = unvisible
                # 15 - CH -
                cell = balans_ws.cell(row=x, column=15)
                cell.value = (
                    f"=IFERROR(VLOOKUP(H{x},SDW!E${sdwrow}:$K${countrow},4,0),0)"
                )
                cell.font = unvisible
                # 16 - CH =
                cell = balans_ws.cell(row=x, column=16)
                cell.value = f'=IFERROR(N{x}-O{x},"")'
                cell.font = unvisible
        print("PASSED BALANS")
